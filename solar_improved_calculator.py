import streamlit as st
import pandas as pd
import math
import requests
from datetime import datetime
import json 
import os 

# Configure page
st.set_page_config(
    page_title="Solar Productive Use Calculator",
    page_icon="☀️",
    layout="wide"
    
)
if 'custom_appliances' not in st.session_state:
    st.session_state.custom_appliances = []
if 'pending_submissions' not in st.session_state:
    st.session_state.pending_submissions = []
if 'show_submission_form' not in st.session_state:
    st.session_state.show_submission_form = False

# Initialize session state
if 'inputs_visible' not in st.session_state:
    st.session_state.inputs_visible = True
if 'calculated' not in st.session_state:
    st.session_state.calculated = False

#File Path 
# File paths for data storage
CUSTOM_APPLIANCES_FILE = "custom_appliances.json"
PENDING_SUBMISSIONS_FILE = "pending_submissions.json"
CUSTOM_APPLIANCES_CSV = "custom_appliances.csv"


# Data loading functions
def load_custom_appliances():
    try:
        if os.path.exists(CUSTOM_APPLIANCES_FILE):
            with open(CUSTOM_APPLIANCES_FILE, 'r') as f:
                return json.load(f)
    except:
        pass
    return []

def load_pending_submissions():
    try:
        if os.path.exists(PENDING_SUBMISSIONS_FILE):
            with open(PENDING_SUBMISSIONS_FILE, 'r') as f:
                return json.load(f)
    except:
        pass
    return []

def save_custom_appliances(data):
    try:
        with open(CUSTOM_APPLIANCES_FILE, 'w') as f:
            json.dump(data, f, indent=2)
    except:
        st.error("Error saving custom appliances")

def save_pending_submissions(data):
    try:
        with open(PENDING_SUBMISSIONS_FILE, 'w') as f:
            json.dump(data, f, indent=2)
    except:
        st.error("Error saving pending submissions")
##new csv custom
def save_custom_appliance_to_csv(appliance_data):
    """Save custom appliance to a standalone CSV file"""
    try:
        # Create DataFrame from appliance data
        df_new = pd.DataFrame([appliance_data])
        
        # Check if CSV exists
        if os.path.exists(CUSTOM_APPLIANCES_CSV):
            # Append to existing CSV
            df_existing = pd.read_csv(CUSTOM_APPLIANCES_CSV)
            df_combined = pd.concat([df_existing, df_new], ignore_index=True)
            df_combined.to_csv(CUSTOM_APPLIANCES_CSV, index=False)
        else:
            # Create new CSV
            df_new.to_csv(CUSTOM_APPLIANCES_CSV, index=False)
        
        return True
    except Exception as e:
        st.error(f"Error saving to CSV: {e}")
        return False


# Minimalist CSS for styling
st.markdown("""
<style>
    :root {
        --primary: #4CAF50;
        --secondary: #2c3e50;
        --accent: #e74c3c;
        --background: #f5f5f5;
        --card: white;
        --text: #333333;
    }
    
    .stApp {
        background-color: var(--background);
    }
    
    .metric-card {
        background-color: var(--card);
        border-radius: 6px;
        padding: 10px;
        margin-bottom: 8px;
        box-shadow: 0 1px 3px rgba(0,0,0,0.05);
        border-left: 2px solid var(--primary);
        height: 80px;
        display: flex;
        flex-direction: column;
        justify-content: center;
    }
    
    .metric-title {
        font-size: 0.75rem;
        color: #000000;
        margin-bottom: 3px;
    }
    
    .metric-value {
        font-size: 1.1rem;
        font-weight: bold;
        color: var(--secondary);
    }
    
    .metric-unit {
        font-size: 0.7rem;
        color: #000000;
    }
    
    .section-title {
        color: var(--secondary);
        border-bottom: 1px solid var(--primary);
        padding-bottom: 5px;
        margin-bottom: 8px;
        font-size: 1rem;
    }
    
    .summary-card {
        background-color: var(--card);
        border-radius: 6px;
        padding: 12px;
        margin: 8px 0;
        box-shadow: 0 1px 3px rgba(0,0,0,0.05);
        font-size: 0.9rem;
    }
    
    .success-box {
        background-color: #e8f5e9;
        border-left: 2px solid #4CAF50;
        padding: 10px;
        border-radius: 0 4px 4px 0;
        margin: 8px 0;
        font-size: 0.9rem;
    }
    
    .warning-box {
        background-color: #fff3e0;
        border-left: 2px solid #ff9800;
        padding: 10px;
        border-radius: 0 4px 4px 0;
        margin: 8px 0;
        font-size: 0.9rem;
    }
    
    .error-box {
        background-color: #ffebee;
        border-left: 2px solid #f44336;
        padding: 10px;
        border-radius: 0 4px 4px 0;
        margin: 8px 0;
        font-size: 0.9rem;
    }
    
    .dataframe {
        border-radius: 6px;
        box-shadow: 0 1px 3px rgba(0,0,0,0.05);
        font-size: 0.85rem;
    }
    
    .stTabs [data-baseweb="tab-list"] {
        gap: 6px;
    }
    
    .stTabs [data-baseweb="tab"] {
        height: 30px;
        padding: 0 12px;
        border-radius: 4px 4px 0 0 !important;
        background-color: #979797 !important;
        font-size: 0.85rem;
        color: #000000 !important;
    }
    
    .stTabs [aria-selected="true"] {
        background-color: var(--primary) !important;
        color: white !important;
    }
    
    .input-section {
        transition: all 0.3s ease;
    }
    
    .collapsed {
        max-height: 0;
        overflow: hidden;
        padding: 0;
        margin: 0;
    }
    
    .minimal-input {
        margin-bottom: 6px;
    }
    
    .compact-expander .streamlit-expanderHeader {
        font-size: 0.95rem;
        padding: 8px 0;
    }
    
    .compact-expander .streamlit-expanderContent {
        padding: 8px 0 0 0;
    }
            /* Change the Modify Inputs button color */
div.stButton > button:first-child {
    background-color: #179C10;
    color: white;
    border: none;
}

div.stButton > button:first-child:hover {
    background-color: #BA0000;
    border: none;
    color: white;
}
            
</style>
""", unsafe_allow_html=True)

# Custom metric card component
def metric_card(title, value, unit="", help_text=None):
    card = f"""
    <div class="metric-card">
        <div class="metric-title">{title}</div>
        <div class="metric-value">{value}</div>
        <div class="metric-unit">{unit}</div>
    </div>
    """
    if help_text:
        with st.expander("ℹ️"):
            st.caption(help_text)
    st.markdown(card, unsafe_allow_html=True)

#custom function

def custom_appliance_submission_form():
    st.markdown("---")
    st.subheader("📝 Submit Custom Appliance to Database")
    
    with st.form("custom_appliance_formnew"):
        col1, col2 = st.columns(2)
        
        with col1:
            name = st.text_input("Appliance Name*", placeholder="e.g., Solar Water Pump 1.5kW")
            power_ac = st.number_input("AC Power Consumption (kW)*", min_value=0.1, value=2.0, step=0.1)
            power_dc = st.number_input("DC Power Consumption (kW)*", min_value=0.1, value=1.5, step=0.1)
            processing_speed = st.number_input("Processing Speed (kg/hour)*", min_value=1, value=100, step=1)
            
        with col2:
            price_usd = st.number_input("Appliance Price (USD)*", min_value=0.0, value=500.0, step=50.0)
            description = st.text_area("Description", placeholder="Brief description of the appliance and its use case")
            contact_email = st.text_input("Your Email (optional)", placeholder="For follow-up questions")
        
        submitted = st.form_submit_button("📤 Submit for Review", type="primary")
        
        if submitted:
            if not name or not power_ac or not power_dc:
                st.error("Please fill in all required fields (*)")
            else:
                submission = {
                    "id": datetime.now().strftime("%Y%m%d%H%M%S"),
                    "name": name,
                    "power_ac": power_ac,
                    "power_dc": power_dc,
                    "processing_speed": processing_speed,
                    "price_usd": price_usd,
                    "description": description,
                    "contact_email": contact_email,
                    "submission_date": datetime.now().isoformat(),
                    "status": "pending"
                }
                
                st.session_state.pending_submissions.append(submission)
                save_pending_submissions(st.session_state.pending_submissions)
                
                st.success("✅ Appliance submitted for review! We'll evaluate it before adding to the database.")
                if st.session_state.get("submitted_success", False):    
                    if st.button("← Back to Calculator"):
                        st.session_state.show_submission_form = False
                        st.session_state.inputs_visible = True
                        st.rerun()

def edit_custom_appliance():
    st.markdown("---")
    st.subheader("✏️ Edit Custom Appliance")

    if not st.session_state.custom_appliances:
        st.info("No custom appliances available to edit.")
        return

    # Select appliance to edit
    appliance_names = [app["name"] for app in st.session_state.custom_appliances]
    selected_name = st.selectbox("Select Appliance to Edit", appliance_names)

    # Find the selected appliance
    appliance = next(app for app in st.session_state.custom_appliances if app["name"] == selected_name)

    # Editable form pre-filled with current values
    with st.form("edit_appliance_form"):
        col1, col2 = st.columns(2)
        with col1:
            name = st.text_input("Appliance Name*", value=appliance["name"])
            power_ac = st.number_input("AC Power (kW)", min_value=0.1, value=float(appliance["power_ac"]), step=0.1)
            power_dc = st.number_input("DC Power (kW)", min_value=0.1, value=float(appliance["power_dc"]), step=0.1)
            processing_speed = st.number_input("Processing Speed (kg/hr)", min_value=1, value=int(appliance["processing_speed"]), step=1)
        with col2:
            price_usd = st.number_input("Price (USD)", min_value=0.0, value=float(appliance["price_usd"]), step=50.0)
            description = st.text_area("Description", value=appliance.get("description", ""))
            contact_email = st.text_input("Contact Email", value=appliance.get("contact_email", ""))

        updated = st.form_submit_button("💾 Update Appliance")

    if updated:
        # Update appliance fields
        appliance.update({
            "name": name,
            "power_ac": power_ac,
            "power_dc": power_dc,
            "processing_speed": processing_speed,
            "price_usd": price_usd,
            "description": description,
            "contact_email": contact_email,
            "last_updated": datetime.now().isoformat()
        })

        # Save back to JSON + CSV
        save_custom_appliances(st.session_state.custom_appliances)
        save_custom_appliance_to_csv(appliance)

        st.success(f"✅ {name} updated successfully!")


# Appliance & system options
# Load custom appliances on startup
if not st.session_state.custom_appliances:
    st.session_state.custom_appliances = load_custom_appliances()
if not st.session_state.pending_submissions:
    st.session_state.pending_submissions = load_pending_submissions()

# Appliance options - combine default and custom appliances
default_appliances = [
    {"name": "Hammer Mill for Flour", "power_ac": 2.0, "power_dc": 2.0, "processing_speed": 100, "price_usd": 600},
    #{"name": "Grade 1 Flour Grinding Hammer Mill 3kW", "power_ac": 3.0, "power_dc": 3.0, "processing_speed": 150, "price_usd": 700},
    
    {"name": "Rice Huller/Polisher", "power_ac": 1.8, "power_dc": 1.3, "processing_speed": 60, "price_usd": 600},
]
custom_appliance_names = [app["name"] for app in st.session_state.custom_appliances]
# Appliance options - combine default and custom appliances
appliances = default_appliances + st.session_state.custom_appliances

# Names only (for dropdowns/UI)
appliance_names = ["Choose one"]+[app["name"] for app in appliances]

# Maps
power_map = {app["name"]: app["power_ac"] for app in appliances}
power_map_dc = {app["name"]: app["power_dc"] for app in appliances}
price_map_usd = {app["name"]: app["price_usd"] for app in appliances}



system_rating = ["Choose one", "AC", "DC"]

# Panel specs
panel_wattage_kw = 0.5  # 500W
panel_cost = 50  

# Maps
power_map = {app["name"]: app["power_ac"] for app in appliances}
power_map_dc = {app["name"]: app["power_dc"] for app in appliances}
price_map_usd = {app["name"]: app["price_usd"] for app in appliances}
processing_speed_map = {app["name"]: app["processing_speed"] for app in appliances}
# Common currencies with sample exchange rates (fallback)
common_currencies = {
    "USD": 1.0,
    "EUR": 0.93,
    "GBP": 0.80,
    "JPY": 154.62,
    "CAD": 1.37,
    "AUD": 1.52,
    "CHF": 0.91,
    "CNY": 7.24,
    "INR": 83.45,
    "BRL": 5.40,
    "MXN": 17.05,
    "ZAR": 18.85,
    "NGN": 1400.0,
    "KES": 133.0,
    "GHS": 13.5,
    "EGP": 47.9,
    "XOF": 610.0
}

# --- GET EXCHANGE RATES WITH FALLBACK ---
@st.cache_data(ttl=3600)  # cache for 1 hour
def get_exchange_rates():
    try:
        # Try multiple API endpoints
        endpoints = [
            "https://api.exchangerate-api.com/v4/latest/USD",
            "https://open.er-api.com/v6/latest/USD"
        ]
        
        for url in endpoints:
            try:
                resp = requests.get(url, timeout=5)
                if resp.status_code == 200:
                    data = resp.json()
                    if "rates" in data:
                        # Merge with our common currencies as fallback
                        rates = data["rates"]
                        for currency, rate in common_currencies.items():
                            if currency not in rates:
                                rates[currency] = rate
                        return rates
            except:
                continue
                
        # If all APIs fail, use our fallback currencies
        st.warning("Could not fetch live exchange rates. Using sample rates.")
        return common_currencies
    except:
        st.warning("Could not fetch exchange rates. Using sample rates.")
        return common_currencies

# --- DETECT USER LOCATION & CURRENCY ---
def get_user_currency():
    try:
        ip_info = requests.get("https://ipapi.co/json/", timeout=3).json()
        country = ip_info.get("country_name", "Unknown")
        currency = ip_info.get("currency", "USD")
        return country, currency
    except:
        return "Unknown", "USD"

# Fetch exchange rates on every rerun
rates = get_exchange_rates()
currencies = sorted(rates.keys())

# ... (previous code remains the same until the currency section)

# Ensure selected currency is valid
if 'selected_currency' not in st.session_state:
    st.session_state.selected_currency = "USD"

# Streamlit UI
st.title("☀️ Solar Productive Use Calculator")

# --- INPUT SECTION ---
if st.session_state.inputs_visible:
    with st.expander("Input Parameters", expanded=True):
        # Create columns for input layout
        col1, col2 = st.columns(2)

        with col1:
            # Currency selection
        
            st.markdown('<div class="section-title">Currency Settings</div>', unsafe_allow_html=True)
            #use_location = st.checkbox(
            use_location = False

            # Get the current selected currency from session state
            selected_currency = st.session_state.selected_currency
            
            if use_location:
                user_country, detected_currency = get_user_currency()
                if detected_currency in currencies:
                    selected_currency = detected_currency
                    st.success(f"Detected location: {user_country} - Using {detected_currency}")
                else:
                    st.warning(f"Detected currency {detected_currency} not supported. Using USD instead.")
                    selected_currency = "USD"
            else:
                # Set the index to USD by finding its position in the currencies list
                usd_index = currencies.index("USD") if "USD" in currencies else 0
                selected_currency = st.selectbox(
                    "Select Currency:", 
                    currencies,
                    index=usd_index
                )
            
            # Update the session state with the selected currency
            st.session_state.selected_currency = selected_currency

          

            rate_new = rates.get(selected_currency, 1)
            if selected_currency != "USD":
                if rates == common_currencies:
                    st.warning(f"⚠️ Using sample exchange rates (1 USD = {rate_new:.2f} {selected_currency}). For accurate results, please verify current rates.")
                else:
                    st.caption(f"💱 Exchange rate used: 1 USD = {rate_new:.2f} {selected_currency}")

            #new changes added above

            
            st.markdown('<div class="section-title">Appliance Details</div>', unsafe_allow_html=True)
            
            # Choice: from database or manual entry
            # Choice: from database or manual entry
            appliance_mode = st.radio("Select Appliance Mode:", ["Pick from Database", "Enter Custom Specs"], horizontal=True)

            if appliance_mode == "Pick from Database":
               
                selected_appliance = st.selectbox(
                    "Productive Use Appliance:", 
                    appliance_names,
                    help="Select the appliance you want to power with solar"
                )
                
                if selected_appliance != "Choose one":
                    custom_appliance = next((app for app in st.session_state.custom_appliances if app["name"] == selected_appliance), None)
                    price_usd = price_map_usd[selected_appliance]
                    processing_speed = processing_speed_map[selected_appliance]
                else:
                    selected_appliance, price_usd, processing_speed = "Choose one", 0, 0

            else:  # Custom Specs
                if st.button("💾 Submit an appliance to database"):
                    st.session_state.inputs_visible = False
                    st.session_state.show_submission_form = True
                    st.rerun()
                    
                    
                custom_name = st.text_input("Appliance Name", value="Custom Mill")
                power = st.number_input("Power Consumption (kW)", min_value=0.1, value=2.0, step=0.1)
                processing_speed = st.number_input("Processing Speed (kg/hour)", min_value=1, value=100, step=1)
                price_usd = st.number_input("Appliance Price (USD)", min_value=0.0, value=500.0, step=50.0)
                selected_appliance = custom_name  
                                                                  
            # System choice always comes AFTER appliance choice
            selected_system = st.selectbox(
                "System Rating:", 
                system_rating,
                help="Select AC or DC system type"
            )

            # Now assign power only once, based on system type
            # Now assign power and appliance specs based on system type and mode
            if appliance_mode == "Pick from Database" and selected_appliance != "Choose one":
                # Check if it's a custom appliance
                custom_appliance = next((app for app in st.session_state.custom_appliances if app["name"] == selected_appliance), None)
                
                if custom_appliance:
                    # Custom appliance found
                    if selected_system == "AC":
                        power = custom_appliance["power_ac"]
                    elif selected_system == "DC":
                        power = custom_appliance["power_dc"]
                    price_usd = custom_appliance["price_usd"]
                    processing_speed = custom_appliance["processing_speed"]
                else:
                    # Default appliance
                    if selected_system == "AC":
                        power = power_map[selected_appliance]
                    elif selected_system == "DC":
                        power = power_map_dc[selected_appliance]
                    price_usd = price_map_usd[selected_appliance]
                    processing_speed = processing_speed_map[selected_appliance]

            

            if selected_appliance == "Hammer Mill for Flour":
                st.markdown("---")
                
                
                # Create a clean container for the image section
                with st.container():
                    
                        # Image with centered alignment and clean styling
                        st.markdown(
                            """
                            <div style="text-align: center; padding: 15px;">
                                <a href="https://productivesolarsolutions.com/uploads/spec/MaizeMill(PoshoMillGrade%202).pdf" 
                                target="_blank">
                                    <img src="https://productivesolarsolutions.com/uploads/products/MaizeMill.png" 
                                        style="max-height: 180px; width: auto; border-radius: 8px; 
                                                border: 1px solid #e0e0e0; padding: 5px;">
                                </a>
                                <div style="margin-top: 8px; font-style: italic; color: #555; font-size: 14px;">
                                    Maize Mill(Posho Mill Grade 2) <br> Can make Grade 2 wholegrain flour from maize and other grains, or Grade 1 maize flour from hulled maize, or coarsely ground animal feed.
                                </div>
                            </div>
                            """, 
                            unsafe_allow_html=True
                        ) #col2
                        

                        
                        st.markdown("""
                        
                        - [Technical Spec Sheet (PDF)](
                        https://productivesolarsolutions.com/uploads/spec/MaizeMill(PoshoMillGrade%202).pdf)
                        - [Product Website](
                        https://productivesolarsolutions.com)
                        """)



            elif selected_appliance == "Grade 1 Flour Grinding Hammer Mill 3kW":
                st.markdown("---")
                
                
                # Create a clean container for the image section
                with st.container():
                    
                        # Image with centered alignment and clean styling
                        st.markdown(
                            """
                            <div style="text-align: center; padding: 15px;">
                                <a href="https://productivesolarsolutions.com/uploads/spec/MaizeMill(PoshoMillGrade%202).pdf" 
                                target="_blank">
                                    <img src="https://www.villageinfrastructure.com/wp-content/uploads/2025/09/new.png" 
                                        style="max-height: 180px; width: auto; border-radius: 8px; 
                                                border: 1px solid #e0e0e0; padding: 5px;">
                                </a>
                                <div style="margin-top: 8px; font-style: italic; color: #555; font-size: 14px;">
                                    Hammer Mill & Rice Huller
                                </div>
                            </div>
                            """, 
                            unsafe_allow_html=True
                        ) #col2
                        

                        
                        st.markdown("""
                        
                        - [Technical Spec Sheet Hammer Mill(PDF)](
                        https://productivesolarsolutions.com/uploads/spec/MaizeMill(PoshoMillGrade%202).pdf)
                        - [Technical Spec Sheet Rice Huller/Polisher (PDF)](
                        https://productivesolarsolutions.com/uploads/spec/MaizeMill(PoshoMillGrade%202).pdf)
                        - [Product Website](
                        https://productivesolarsolutions.com)
                        """) 
            
            elif selected_appliance == "Rice Huller/Polisher":
                st.markdown("---")

                
                # Create a clean container for the image section
                with st.container():
                    
                    
                    
                        # Image with centered alignment and clean styling
                        st.markdown(
                            """
                            <div style="text-align: center; padding: 15px;">
                                <a href="https://productivesolarsolutions.com/uploads/spec/MaizeMill(PoshoMillGrade%202).pdf" 
                                target="_blank">
                                    <img src="https://productivesolarsolutions.com/uploads/products/RiceMill-MaizeHuller.png" 
                                        style="max-height: 180px; width: auto; border-radius: 8px; 
                                                border: 1px solid #e0e0e0; padding: 5px;">
                                </a>
                                <div style="margin-top: 8px; font-style: italic; color: #555; font-size: 14px;">
                                    Rice Mill - Maize Huller <br> Hulls and polishes rice to various levels of whiteness. Can also hull maize for making Grade 1 flour when combined with a hammer mill.
                                </div>
                            </div>
                            """, 
                            unsafe_allow_html=True
                        )
                        
                
                        
                        st.markdown("""
                        
                        - [Technical Spec Sheet (PDF)](
                        https://productivesolarsolutions.com/uploads/spec/MaizeMill(PoshoMillGrade%202).pdf)
                        - [Product Website](
                        https://productivesolarsolutions.com)
                        """)     

            

            runtime_per_day = st.slider(
                "Runtime Per Day (hrs)", 
                min_value=1.0, 
                max_value=24.0,
                value=4.0, 
                step=0.5,
                help="Daily operating hours of the appliance"
            )
            
            operating_days = st.slider(
                "Operating Days per Year", 
                min_value=1, 
                max_value=365,
                value=250,
                help="Number of days per year the business will operate"
            )
            
            #input selection
            if selected_currency == "USD":
                income_per_kg = st.number_input(
                    "Income per kg (USD)", 
                    min_value=0.0, 
                    value=round(5/140, 3),
                    step=0.001,
                    format="%.3f",
                    help="Revenue generated per kg of processed material"
                )
                income_per_kg_usd = income_per_kg  # Already in USD
            else:
                # Income per kg in selected currency
                income_per_kg_local = st.number_input(
                    f"Income per kg ({selected_currency})", 
                    min_value=0.0, 
                    value=round((5/140) * rates.get(selected_currency, 1.0), 3),  # Convert default value
                    step=0.001,
                    format="%.3f",
                    help=f"Revenue generated per kg of processed material in {selected_currency}"
                )
                # Convert back to USD for calculations
                exchange_rate = rates.get(selected_currency, 1.0)
                income_per_kg_usd = income_per_kg_local / exchange_rate

            # Use the USD value for calculations
            income_per_kg = income_per_kg_usd

        with col2:
            st.markdown('<div class="section-title">Solar System Details</div>', unsafe_allow_html=True)
            sun_hours = st.slider(
                "Sun Hours Per Day (hrs)", 
                min_value=1.0, 
                max_value=12.0,
                value=4.0, 
                step=0.5,
                help="Average daily peak sun hours at your location"
            )
            
            system_efficiency = st.slider(
                "System Efficiency (%)", 
                min_value=1, 
                max_value=100, 
                value=80,
                help="Overall efficiency of the solar system"
            )
            
            # Changed from slider to number input for battery storage
            battery_hours = st.number_input(
                "Battery Storage (hrs)", 
                min_value=0, 
                max_value=24,
                value=1,
                help="Hours of battery backup required"
            )
            
            if selected_currency == "USD":
                daily_operating_cost = st.number_input(
                    "Daily Operating Cost (USD)", 
                    value=10.0, 
                    step=1.0,
                    help="Daily expenses like labor, rent, etc."
                )
                daily_operating_cost_usd = daily_operating_cost  # Already in USD
            else:
                # Daily operating cost in selected currency
                daily_operating_cost_local = st.number_input(
                    f"Daily Operating Cost ({selected_currency})", 
                    value=10.0 * rates.get(selected_currency, 1.0),  # Convert default value
                    step=1.0,
                    help=f"Daily expenses like labor, rent, etc. in {selected_currency}"
                )
                # Convert back to USD for calculations
                exchange_rate = rates.get(selected_currency, 1.0)
                daily_operating_cost_usd = daily_operating_cost_local / exchange_rate

            # Use the USD value for calculations
            daily_operating_cost = daily_operating_cost_usd

        # Financial inputs
        st.markdown('<div class="section-title">Financing Options</div>', unsafe_allow_html=True)
        col3, col4, col5, col6 = st.columns(4)

        with col3:
            loan_term_years = st.slider(
                "Loan Term (Years)", 
                min_value=1,
                max_value=10,
                value=3, 
                step=1,
                help="Duration of the loan repayment period"
            )

        with col4:
            interest_rate = st.slider(
                "Interest Rate (p.a. %)", 
                min_value=0.0,
                max_value=30.0,
                value=15.0, 
                step=0.5,
                help="Annual interest rate percentage for the loan"
            ) / 100

        with col5:
            # Changed to percentage slider for deposit
            deposit_percentage = st.slider(
                "Deposit (% of total cost)", 
                min_value=0, 
                max_value=100, 
                value=0,
                step=5,
                help="Percentage of the total installed cost as deposit"
            )

        with col6:
            install_increase = st.slider(
                "Import & Installation Cost Increase (%)", 
                min_value=0, 
                max_value=500,
                value=100, 
                step=10,
                help="Additional percentage cost for importing and installation"
            )
            install_multiplier = 1 + (install_increase / 100)

        # Subsidy input
        st.markdown('<div class="section-title">Subsidy & Grant Options</div>', unsafe_allow_html=True)
        subsidy_percentage = st.slider(
            "Subsidy Percentage (%)",
            min_value=0,
            max_value=100,
            value=0,
            step=5,
            help="Percentage of the total installed cost covered by subsidies"
        )

        

    calculate_btn = st.button(
        "🚀 Calculate System Requirements", 
        use_container_width=True,
        type="primary"
    )

    if calculate_btn:
        if selected_appliance != "Choose one" and selected_system != "Choose one":
            st.session_state.inputs_visible = False
            st.session_state.calculated = True
            st.session_state.selected_appliance = selected_appliance
            st.session_state.selected_system = selected_system
            st.session_state.runtime_per_day = runtime_per_day
            st.session_state.operating_days = operating_days
            st.session_state.income_per_kg = income_per_kg
            st.session_state.sun_hours = sun_hours
            st.session_state.system_efficiency = system_efficiency
            st.session_state.battery_hours = battery_hours
            st.session_state.daily_operating_cost = daily_operating_cost
            st.session_state.loan_term_years = loan_term_years
            st.session_state.interest_rate = interest_rate
            st.session_state.deposit_percentage = deposit_percentage
            st.session_state.install_multiplier = install_multiplier
            st.session_state.subsidy_percentage = subsidy_percentage
            st.session_state.selected_currency = selected_currency
            
            # Store appliance-specific values
            if appliance_mode == "Pick from Database":
                if selected_system == "AC":
                    st.session_state.power = power_map[selected_appliance]
                elif selected_system == "DC":
                    st.session_state.power = power_map_dc[selected_appliance]
                st.session_state.price_usd = price_map_usd[selected_appliance]
                st.session_state.processing_speed = processing_speed_map[selected_appliance]
            else:
                st.session_state.power = power
                st.session_state.price_usd = price_usd
                st.session_state.processing_speed = processing_speed
                
            st.rerun()
        else:
            st.error("⚠️ Please select both an Appliance and System type before calculating.")

#submission triggering
if st.session_state.get('show_submission_form', False):
    custom_appliance_submission_form()
    if st.button("← Back to Calculator"):
        st.session_state.show_submission_form = False
        st.session_state.inputs_visible = True
        st.rerun()

# --- RESULTS SECTION ---
if not st.session_state.inputs_visible and st.session_state.get("calculated", False):
    # Get values from session state
    selected_appliance = st.session_state.selected_appliance
    selected_system = st.session_state.selected_system
    runtime_per_day = st.session_state.runtime_per_day
    operating_days = st.session_state.operating_days
    income_per_kg = st.session_state.income_per_kg
    sun_hours = st.session_state.sun_hours
    system_efficiency = st.session_state.system_efficiency
    battery_hours = st.session_state.battery_hours
    daily_operating_cost = st.session_state.daily_operating_cost
    loan_term_years = st.session_state.loan_term_years
    interest_rate = st.session_state.interest_rate
    deposit_percentage = st.session_state.deposit_percentage
    install_multiplier = st.session_state.install_multiplier
    subsidy_percentage = st.session_state.subsidy_percentage
    selected_currency = st.session_state.selected_currency
    
    # Get appliance-specific values
    power = st.session_state.power
    price_usd = st.session_state.price_usd
    processing_speed = st.session_state.processing_speed

    # Get exchange rate but don't convert yet
    rate = rates.get(selected_currency, 1)

    # Calculations - all in USD
    specific_efficiency = processing_speed / power
    energy_required_per_day = runtime_per_day * power
    energy_production = energy_required_per_day / (system_efficiency / 100)
    production_per_day = specific_efficiency * energy_required_per_day
    income_per_hour = income_per_kg * processing_speed
    income_per_day = income_per_kg * production_per_day
    gross_income_per_year = income_per_day * operating_days
    net_income_per_day = income_per_day - daily_operating_cost
    panel_energy_per_day = panel_wattage_kw * sun_hours 
    panels_required = math.ceil(energy_production / panel_energy_per_day)
    solar_panel_cost = panels_required * panel_cost  # USD
    recommended_solar_size = math.ceil((energy_production / sun_hours) * 2) / 2
    battery_capacity = recommended_solar_size * battery_hours

    # Initialize both costs to 0 in USD
    inverter_cost = 0
    controller_cost = 0

    if selected_system == "AC":
        inverter_cost = recommended_solar_size * 100   # USD
    elif selected_system == "DC":
        controller_cost = recommended_solar_size * 50  # USD
        
    # Battery cost in USD
    battery_cost = battery_capacity * 300  # USD

    # Cost calculations in USD
    fob_subtotal_usd = price_usd + solar_panel_cost + inverter_cost + controller_cost + battery_cost
    total_with_import_usd = fob_subtotal_usd * install_multiplier
    
    # Apply subsidy
    subsidy_amount = total_with_import_usd * (subsidy_percentage / 100)
    total_after_subsidy = total_with_import_usd - subsidy_amount
    
    # Calculate deposit amount based on percentage
    deposit_amount = total_after_subsidy * (deposit_percentage / 100)
    loan_principal_usd = total_after_subsidy - deposit_amount

    # Loan calculations in USD
    months = loan_term_years * 12
    monthly_rate = interest_rate / 12

    if monthly_rate > 0 and loan_principal_usd > 0:
        monthly_repayment_usd = (loan_principal_usd * monthly_rate) / (1 - (1 + monthly_rate)**(-months))
    else:
        monthly_repayment_usd = 0

    total_repayment_usd = months * monthly_repayment_usd
    total_interest_paid_usd = total_repayment_usd - loan_principal_usd
    annual_repayment_usd = monthly_repayment_usd * 12
    daily_repayment_usd = annual_repayment_usd / 365

    if income_per_day > 0:
        repayment_percentage = (daily_repayment_usd / income_per_day) * 100
    else:
        repayment_percentage = 0
    
    # New metric: % of Daily Net Revenue Used for Repayment
    if net_income_per_day > 0:
        net_revenue_repayment_percentage = (daily_repayment_usd / net_income_per_day) * 100
    else:
        net_revenue_repayment_percentage = 0
    
    # Business viability - fixed logic
    # Business is viable if no loan is needed (deposit is 100%) OR net income covers repayments
    if deposit_percentage == 100 or (net_income_per_day > 0 and daily_repayment_usd > 0 and net_income_per_day >= daily_repayment_usd):
        viable_business = True
        viability_text = "Yes ✅" 
        viability_class = "success-box"
    else:
        viable_business = False
        viability_text = "No ❌"
        viability_class = "error-box"

    # Display results in tabs
    tab1, tab2, tab3, tab4, tab5, tab6 = st.tabs(["📊 Overview", "💵 Financials", "⚡ Technical", "📈 Viability", "📊Chart", "Export Report"])

    with tab1:
        st.subheader("Key Metrics")
        
        col1, col2, col3, col4 = st.columns(4)
        with col1:
            metric_card(
                "Solar Size", 
                f"{recommended_solar_size}", 
                "kWp",
                "Total solar capacity needed"
            )
        with col2:
            metric_card(
                "Panels Required", 
                f"{panels_required}", 
                "panels",
                "Number of solar panels needed"
            )
        with col3:
            metric_card(
                "Daily Production", 
                f"{round(production_per_day, 1)}", 
                "kg/day",
                "Estimated daily processing output"
            )
        with col4:
            metric_card(
                "Daily Net Income", 
                f"{round(net_income_per_day * rate, 1)}", 
                selected_currency,
                "Income after operating costs"
            )
        
        st.markdown("---")
        st.subheader("System Overview")

        st.markdown(f"""
        <div class="summary-card">
            <p><b>Machine Details:</b> {selected_appliance} ({power}kW {selected_system} system)</p>
            <p><b>Daily Operation:</b> {runtime_per_day} hours/day, {operating_days} days/year</p>
            <p><b>Solar Requirements:</b> {panels_required} x 500W panels ({recommended_solar_size} kWp system)</p>
            <p><b>Battery Storage:</b> {battery_capacity} kWh ({battery_hours} hours backup)</p>
            <p><b>Location:</b> {sun_hours} peak sun hours per day</p>
            <p><b>System Efficiency:</b> {system_efficiency}%</p>
        </div>
        """, unsafe_allow_html=True)

        # --- Organized Image + Tech Spec Link for Maize Mill ---
        if selected_appliance == "Hammer Mill for Flour":
            st.markdown("---")
            st.subheader("Technical Specifications")
            
            # Create a clean container for the image section
            with st.container():
                col1, col2 = st.columns([1, 2])
                
                with col1:
                    # Image with centered alignment and clean styling
                    st.markdown(
                        """
                        <div style="text-align: center; padding: 15px;">
                            <a href="https://productivesolarsolutions.com/uploads/spec/MaizeMill(PoshoMillGrade%202).pdf" 
                            target="_blank">
                                <img src="https://productivesolarsolutions.com/uploads/products/MaizeMill.png" 
                                    style="max-height: 180px; width: auto; border-radius: 8px; 
                                            border: 1px solid #e0e0e0; padding: 5px;">
                            </a>
                            <div style="margin-top: 8px; font-style: italic; color: #555; font-size: 14px;">
                                Maize Mill(Posho Mill Grade 2)
                            </div>
                        </div>
                        """, 
                        unsafe_allow_html=True
                    )
                    
                with col2:
                    # Conditional specifications based on AC/DC selection
                    if selected_system == "DC":
                        st.markdown("""
                        **Power Parameters: DC Powered Hammer Mill**
                        - **Motor Type:** Brushless DC motor
                        - **Power:** 1.3 KW / 1.74 Hp
                        - **Voltage:** 48V
                        - **Controller:** 2 KW DC Motor controller
                        - **Recommended Solar:** 1KW solar panels
                        - **Battery Options:**
                            - 4 x 100Ah Lead-Acid batteries OR
                            - 1 x 200Ah Lithium-Ion battery
                        """)
                    else:  # AC system
                        st.markdown("""
                        **Power Parameters: AC Grid Power**
                        - **Motor Type:** AC Electric motor
                        - **Power:** 2.2 KW / 3 Hp
                        - **Voltage:** 220/240 V
                        - **Starter:** 7-10A DOL Starter
                        - **Protection:** Industrial socket and plug
                        - **Power Source:** Grid connection required
                        """)
                    
                    st.markdown("""
                    **Documentation:**
                    - [Technical Spec Sheet (PDF)](
                    https://productivesolarsolutions.com/uploads/spec/MaizeMill(PoshoMillGrade%202).pdf)
                    - [Product Website](
                    https://productivesolarsolutions.com)
                    """)
            
            # Clean caption below the container
            st.caption("Click the image or links above to view detailed technical specifications")

        # --- Organized Image + Tech Spec Link for Rice Mill ---
        elif selected_appliance == "Rice Huller/Polisher":
            st.markdown("---")
            st.subheader("Technical Specifications")
            
            # Create a clean container for the image section
            with st.container():
                col1, col2 = st.columns([1, 2])
                
                with col1:
                    # Image with centered alignment and clean styling
                    st.markdown(
                        """
                        <div style="text-align: center; padding: 15px;">
                            <a href="https://productivesolarsolutions.com/uploads/spec/MaizeMill(PoshoMillGrade%202).pdf" 
                            target="_blank">
                                <img src="https://productivesolarsolutions.com/uploads/products/RiceMill-MaizeHuller.png" 
                                    style="max-height: 180px; width: auto; border-radius: 8px; 
                                            border: 1px solid #e0e0e0; padding: 5px;">
                            </a>
                            <div style="margin-top: 8px; font-style: italic; color: #555; font-size: 14px;">
                                Rice Mill - Maize Huller
                            </div>
                        </div>
                        """, 
                        unsafe_allow_html=True
                    )
                    
                with col2:
                    # Conditional specifications based on AC/DC selection
                    if selected_system == "DC":
                        st.markdown("""
                        **Power Parameters: DC Power**
                        - **Motor Type:** Brushless DC motor
                        - **Power:** 1.3 KW / 1.74 Hp
                        - **Voltage:** 48V
                        - **Controller:** 2 KW DC Motor controller
                        - **Recommended Solar:** 1KW solar panels
                        - **Battery Options:**
                            - 4 x 100Ah Lead-Acid batteries OR
                            - 1 x 200Ah Lithium-Ion battery
                        """)
                    else:  # AC system
                        st.markdown("""
                        **Power Parameters: AC Grid Power**
                        - **Motor Type:** AC Electric motor
                        - **Power:** 1.8 KW / 2.4 Hp
                        - **Voltage:** 220/240 V
                        - **Starter:** 7-10A DOL Starter
                        - **Protection:** Industrial socket for power protection
                        - **Power Source:** Grid connection required
                        """)
                    
                    st.markdown("""
                    **Documentation:**
                    - [Technical Spec Sheet (PDF)](
                    https://productivesolarsolutions.com/uploads/spec/MaizeMill(PoshoMillGrade%202).pdf)
                    - [Product Website](
                    https://productivesolarsolutions.com)
                    """)
            
            # Clean caption below the container
            st.caption("Click the image or links above to view detailed technical specifications")    
    
    
    with tab2:
        st.subheader("Cost Breakdown")
        
        col1, col2 = st.columns(2)
        with col1:
            metric_card(
                "Machine Cost", 
                f"{round(price_usd * rate, 1)}", 
                selected_currency
            )
            metric_card(
                "Solar Panel Cost", 
                f"{round(solar_panel_cost * rate, 1)}", 
                selected_currency
            )
            metric_card(
                "Battery Cost", 
                f"{round(battery_cost * rate, 1)}", 
                selected_currency
            )
            
        with col2:
            if selected_system == "AC":
                metric_card(
                    "Inverter Cost", 
                    f"{round(inverter_cost * rate, 1)}", 
                    selected_currency
                )
            else:
                metric_card(
                    "Controller Cost", 
                    f"{round(controller_cost * rate, 1)}", 
                    selected_currency
                )
            metric_card(
                "Import & Installation", 
                f"{round((fob_subtotal_usd * (install_multiplier - 1)) * rate, 1)}", 
                selected_currency
            )
        
        st.markdown("---")
        st.subheader("Total Costs")
        
        col3, col4, col5 = st.columns(3)
        with col3:
            metric_card(
                "FOB Subtotal", 
                f"{round(fob_subtotal_usd * rate, 1)}", 
                selected_currency
            )
        with col4:
            metric_card(
                "Installed Cost", 
                f"{round(total_with_import_usd * rate, 1)}", 
                selected_currency
            )
        with col5:
            metric_card(
                "Subsidy Amount", 
                f"{round(subsidy_amount * rate, 1)}", 
                selected_currency
            )
        
        st.markdown("---")
        st.subheader("Final Costs After Subsidy")
        
        col6, col7, col8 = st.columns(3)
        with col6:
            metric_card(
                "Total After Subsidy", 
                f"{round(total_after_subsidy * rate, 1)}", 
                selected_currency
            )
        with col7:
            metric_card(
                "Deposit Amount", 
                f"{round(deposit_amount * rate, 1)}", 
                selected_currency
            )
        with col8:
            metric_card(
                "Loan Amount", 
                f"{round(loan_principal_usd * rate, 1)}", 
                selected_currency
            )
        
        st.markdown("---")
        st.subheader("Loan Details")
        
        col75, col9, col10, col11  = st.columns(4)
        with col75:
            metric_card(
                "Annual Repayment",
                f"{round(annual_repayment_usd*rate,1)}",
                selected_currency
            )
        with col9:
            metric_card(
                "Monthly Repayment", 
                f"{round(monthly_repayment_usd * rate, 1)}", 
                selected_currency
            )
        with col10:
            metric_card(
                "Daily Repayment", 
                f"{round(daily_repayment_usd * rate, 1)}", 
                selected_currency
            )
        with col11:
            metric_card(
                "Total Interest", 
                f"{round(total_interest_paid_usd * rate, 1)}", 
                selected_currency
            )
        
        st.markdown("---")
        st.subheader("Repayment Analysis")
        
        col12, col13 = st.columns(2)
        with col12:
            metric_card(
                "% of Gross Revenue", 
                f"{round(repayment_percentage, 1)}", 
                "%"
            )
        with col13:
            metric_card(
                "% of Net Revenue", 
                f"{round(net_revenue_repayment_percentage, 1)}", 
                "%"
            )

    with tab3:
        st.subheader("Technical Specifications")
        
        col1, col2, col3 = st.columns(3)
        with col1:
            metric_card(
                "Machine Power", 
                f"{power}", 
                "kW"
            )
        with col2:
            metric_card(
                "Daily Energy Required", 
                f"{round(energy_required_per_day, 1)}", 
                "kWh/day"
            )
        with col3:
            metric_card(
                "Daily Energy Production", 
                f"{round(energy_production, 1)}", 
                "kWh/day"
            )
        
        st.markdown("---")
        st.subheader("Performance Metrics")
        
        col4, col5, col6 = st.columns(3)
        with col4:
            metric_card(
                "Processing Speed", 
                f"{processing_speed}", 
                "kg/hour"
            )
        with col5:
            metric_card(
                "Specific Efficiency", 
                f"{round(specific_efficiency, 2)}", 
                "kg/kWh"
            )
        with col6:
            metric_card(
                "Battery Backup", 
                f"{battery_hours}", 
                "hours"
            )
        
        st.markdown("---")
        st.subheader("Solar System Details")
        
        col7, col8, col9 = st.columns(3)
        with col7:
            metric_card(
                "Panel Wattage", 
                f"{panel_wattage_kw*1000}", 
                "W"
            )
        with col8:
            metric_card(
                "Sun Hours", 
                f"{sun_hours}", 
                "hours/day"
            )
        with col9:
            metric_card(
                "System Efficiency", 
                f"{system_efficiency}", 
                "%"
            )
        
        st.markdown("---")
        st.subheader("Detailed Calculations")
        
        df_tech = pd.DataFrame([{
            "Parameter": "Machine Power",
            "Value": f"{power}",
            "Unit": "kW"
        }, {
            "Parameter": "Daily Runtime",
            "Value": runtime_per_day,
            "Unit": "hours"
        }, {
            "Parameter": "Energy Required",
            "Value": round(energy_required_per_day, 2),
            "Unit": "kWh/day"
        }, {
            "Parameter": "System Efficiency",
            "Value": system_efficiency,
            "Unit": "%"
        }, {
            "Parameter": "Energy Production Needed",
            "Value": round(energy_production, 2),
            "Unit": "kWh/day"
        }, {
            "Parameter": "Sun Hours Available",
            "Value": sun_hours,
            "Unit": "hours"
        }, {
            "Parameter": "Solar System Size",
            "Value": recommended_solar_size,
            "Unit": "kWp"
        }, {
            "Parameter": "Panel Wattage",
            "Value": panel_wattage_kw*1000,
            "Unit": "W"
        }, {
            "Parameter": "Panels Required",
            "Value": panels_required,
            "Unit": "panels"
        }, {
            "Parameter": "Production Rate",
            "Value": processing_speed,
            "Unit": "kg/hour"
        }, {
            "Parameter": "Daily Production",
            "Value": round(production_per_day, 2),
            "Unit": "kg/day"
        }, {
            "Parameter": "Battery Storage",
            "Value": battery_capacity,
            "Unit": "kWh"
        }])
        
        st.dataframe(
            df_tech, 
            hide_index=True, 
            use_container_width=True,
            column_config={
                "Parameter": st.column_config.Column(width="medium"),
                "Value": st.column_config.Column(width="small"),
                "Unit": st.column_config.Column(width="small")
            }
        )

    with tab4:
        st.subheader("Business Viability Analysis")
        
        st.markdown(f"""
        <div class="{viability_class}">
            <h3>Viable Business? {viability_text}</h3>
            <p>Net Income: {round(net_income_per_day * rate, 1)} {selected_currency}/day</p>
            <p>Loan Repayment: {round(daily_repayment_usd * rate, 1)} {selected_currency}/day</p>
        </div>
        """, unsafe_allow_html=True)
        
        st.markdown("---")
        st.subheader("Income vs. Repayments")
        
        col1, col2, col3 = st.columns(3)
        with col1:
            metric_card(
                "Daily Gross Income", 
                f"{round(income_per_day * rate, 1)}", 
                selected_currency
            )
        with col2:
            metric_card(
                "Daily Operating Cost", 
                f"{round(daily_operating_cost * rate, 1)}", 
                selected_currency
            )
        with col3:
            metric_card(
                "Daily Net Income", 
                f"{round(net_income_per_day * rate, 1)}", 
                selected_currency
            )
        
        st.markdown("---")
        st.subheader("Loan Repayment Details")
        
        col4, col5, col6 = st.columns(3)
        with col4:
            metric_card(
                "Daily Loan Repayment", 
                f"{round(daily_repayment_usd * rate, 1)}", 
                selected_currency
            )
        with col5:
            metric_card(
                "% of Gross Revenue", 
                f"{round(repayment_percentage, 1)}", 
                "%"
            )
        with col6:
            metric_card(
                "% of Net Revenue", 
                f"{round(net_revenue_repayment_percentage, 1)}", 
                "%"
            )
        
        st.markdown("---")
        st.subheader("Financial Ratios")
        
        col7, col8 = st.columns(2)
        with col7:
            if viable_business and net_income_per_day > daily_repayment_usd:
                surplus = net_income_per_day - daily_repayment_usd
                metric_card(
                    "Daily Surplus", 
                    f"{round(surplus * rate, 1)}", 
                    selected_currency
                )
        with col8:
            metric_card(
                "Annual Net Profit", 
                f"{round(net_income_per_day * operating_days * rate, 1)}", 
                selected_currency
            )
        
        st.markdown("---")
        st.subheader("Payback Analysis")
        
        if viable_business:
            payback_years = total_after_subsidy / (net_income_per_day * operating_days)
            st.markdown(f"""
            <div class="summary-card">
                <p><b>Your Total Investment:</b> {round(total_after_subsidy * rate, 1)} {selected_currency}</p>
                <p><b>Annual Net Profit:</b> {round(net_income_per_day * operating_days * rate, 1)} {selected_currency}</p>
                <p><b>Simple Payback Period:</b> {round(payback_years, 1)} years</p>
            </div>
            """, unsafe_allow_html=True)
        else:
            st.warning("Payback analysis not available - business is not viable")

    
    # Add a button to show inputs again
    if st.button("↻ Modify Inputs", use_container_width=True):
        st.session_state.inputs_visible = True
        st.session_state.calculated = False
        st.rerun()

    # Show exchange rate disclaimer if using fallback rates
    if selected_currency != "USD":
        if rates == common_currencies:
            st.warning(f"⚠️ Using sample exchange rates (1 USD = {rate:.2f} {selected_currency}). For accurate results, please verify current rates.")
        else:
            st.caption(f"💱 Exchange rate used: 1 USD = {rate:.2f} {selected_currency}")
    with tab5:
        st.subheader("Financial Projection Timeline")
        
        # Calculate key metrics for cashflow projection
        revenue_per_day = income_per_day
        operating_cost_per_day = daily_operating_cost
        loan_repayment_per_day = daily_repayment_usd
        net_income_per_day = revenue_per_day - operating_cost_per_day
        
        # Convert to monthly values
        revenue_per_month = revenue_per_day * 30
        operating_cost_per_month = operating_cost_per_day * 30
        loan_repayment_per_month = loan_repayment_per_day * 30
        net_income_per_month = net_income_per_day * 30
        
        # Define projection periods in months
        months = list(range(1, 121))  # 10 years (120 months)
        
        # Initialize arrays for tracking
        remaining_finance = []
        cumulative_revenue = []
        cumulative_net_profit = []
        break_even_crossed = False
        break_even_month = None
        
        # Initial investment (negative value)
        current_balance = -total_after_subsidy  # Initial investment
        loan_remaining = loan_principal_usd
        
        for month in months:
            # Monthly cash flow
            monthly_revenue = revenue_per_month
            monthly_opex = operating_cost_per_month
            
            # Loan repayment for this month
            if loan_remaining > 0:
                monthly_repayment = min(loan_remaining, loan_repayment_per_month)
                loan_remaining -= monthly_repayment
            else:
                monthly_repayment = 0
            
            # Net cash flow for the month
            monthly_net_cash = monthly_revenue - monthly_opex - monthly_repayment
            
            # Update cumulative values
            current_balance += monthly_net_cash
            
            # Track break-even point
            if not break_even_crossed and current_balance >= 0:
                break_even_crossed = True
                break_even_month = month
            
            remaining_finance.append(current_balance * rate)
            cumulative_revenue.append(monthly_revenue * month * rate)
            cumulative_net_profit.append((monthly_revenue - monthly_opex) * month * rate)
        
        # Plot - Combined bar and line chart
        import plotly.graph_objects as go

        # X-axis in years
        years = [m/12 for m in months]

        # Create interactive figure
        fig = go.Figure()

        # Finance remaining as bars
        finance_colors = []
        for finance in remaining_finance:
            if finance >= 0:  # After break-even
                finance_colors.append("green")
            else:  # Before break-even
                finance_colors.append("red")

        # Finance remaining as bars with conditional coloring
        fig.add_trace(go.Bar(
            x=years, 
            y=remaining_finance, 
            name="Finance Remaining",
            marker_color=finance_colors,  # Now it's an array of colors
            opacity=0.7,
            hovertemplate="Year: %{x:.1f}<br>Remaining: %{y:,.0f}"
        ))

        # Revenue line
        fig.add_trace(go.Scatter(
            x=years, y=cumulative_revenue,
            mode="lines+markers",
            name="Cumulative Revenue",
            line=dict(color="blue", width=2),
            hovertemplate="Year: %{x:.1f}<br>Revenue: %{y:,.0f}"
        ))

        # Net Profit line
        fig.add_trace(go.Scatter(
            x=years, y=cumulative_net_profit,
            mode="lines+markers",
            name="Cumulative Net Profit",
            line=dict(color="green", width=2),
            hovertemplate="Year: %{x:.1f}<br>Net Profit: %{y:,.0f}"
        ))

        # Add break-even line
        fig.add_hline(y=0, line_dash="dash", line_color="yellowgreen", 
                    annotation_text="Break-even", annotation_position="bottom right")

        if break_even_crossed:
            break_even_year = break_even_month / 12
            fig.add_vline(x=break_even_year, line_dash="dot", line_color="orange", 
                        annotation_text=f"Break-even {break_even_year:.1f} yrs", 
                        annotation_position="top")
            fig.add_trace(go.Scatter(
                x=[break_even_year], y=[0],
                mode="markers", name="Break-even Point",
                marker=dict(size=12, color="orange", line=dict(width=1, color="black"))
            ))

        # Layout styling
        fig.update_layout(
            title="Financial Projection: Revenue, Net Profit & Remaining Finance",
            xaxis_title="Years",
            yaxis_title=f"Amount ({selected_currency})",
            template="plotly_white",
            legend=dict(x=0.01, y=0.99, bgcolor="rgba(255,255,255,0.7)"),
            height=600,
            
            # Disable zoom / panning
            xaxis=dict(fixedrange=True),
            yaxis=dict(fixedrange=True),
        )

        # Show in Streamlit
        st.plotly_chart(fig, use_container_width=True)

        
        # Add summary statistics
        st.markdown("---")
        st.subheader("Financial Projection Summary")
        
        col1, col2, col3, col4 = st.columns(4)
        with col1:
            metric_card(
                "Initial Investment", 
                f"{round(total_after_subsidy * rate, 1):,}", 
                selected_currency
            )
        with col2:
            if break_even_crossed:
                metric_card(
                    "Break-even Point", 
                    f"{break_even_month/12:.1f}", 
                    "years"
                )
            else:
                metric_card(
                    "Break-even Point", 
                    ">10", 
                    "years"
                )
        with col3:
            metric_card(
                "10-Year Net Position", 
                f"{round(remaining_finance[-1], 1):,}", 
                selected_currency
            )
        with col4:
            metric_card(
                "Monthly Net Cash Flow", 
                f"{round((net_income_per_month - loan_repayment_per_month) * rate, 1):,}", 
                selected_currency
            )
        
        # Detailed breakdown table for key milestones
        milestone_months = [6, 12, 24, 36, 48, 60, 72, 84, 96, 108, 120]  # 6mo, 1yr, 2yr, 3yr, 5yr, 10yr
        milestone_data = []
        
        for milestone in milestone_months:
            if milestone <= len(months):
                idx = milestone - 1
                milestone_data.append({
                    "Period": f"{milestone} months ({milestone/12:.1f} years)",
                    "Net Position": f"{round(remaining_finance[idx], 1):,} {selected_currency}",
                    "Cumulative Revenue": f"{round(cumulative_revenue[idx], 1):,} {selected_currency}",
                    "Cumulative Net Profit": f"{round(cumulative_net_profit[idx], 1):,} {selected_currency}"
                })
        
        st.markdown("---")
        st.subheader("Key Milestones")
        st.dataframe(pd.DataFrame(milestone_data), hide_index=True, use_container_width=True)
        
    with tab6:
        st.divider()
        st.subheader("📊 Comprehensive Export Report")
        install_increase = (install_multiplier - 1) * 100
        # Payback calculation
        payback_years_new = total_after_subsidy / (net_income_per_day * operating_days) if net_income_per_day > 0 else float('inf')
        
        # Create comprehensive report data with ALL inputs and outputs
        report_data = {
            # Project Info
            "Project": "Solar Productive Use Analysis",
            "Date": datetime.now().strftime("%Y-%m-%d %H:%M"),
            "Currency": selected_currency,
            "Exchange Rate": f"1 USD = {rate:.2f} {selected_currency}",
            
            # INPUTS - Appliance Details
            "Appliance": selected_appliance,
            "System Type": selected_system,
            "Power Consumption": f"{power} kW",
            "Processing Speed": f"{processing_speed} kg/hour",
            "Appliance Price": f"${price_usd} USD",
            "Daily Runtime": f"{runtime_per_day} hours",
            "Operating Days/Year": operating_days,
            
            # INPUTS - Solar System
            "Sun Hours/Day": f"{sun_hours} hours",
            "System Efficiency": f"{system_efficiency}%",
            "Battery Storage": f"{battery_hours} hours",
            
            # INPUTS - Financial
            "Income per kg": f"{selected_currency} {round(income_per_kg * rate, 3)}",
            "Daily Operating Cost": f"{selected_currency} {round(daily_operating_cost * rate, 1)}",
            "Loan Term": f"{loan_term_years} years",
            "Interest Rate": f"{interest_rate * 100}%",
            "Deposit Percentage": f"{deposit_percentage}%",
            "Import/Installation Cost Increase": f"{install_increase}%",
            "Subsidy Percentage": f"{subsidy_percentage}%",
            
            # OUTPUTS - Technical Specifications
            "Recommended Solar Size": f"{recommended_solar_size} kWp",
            "Panels Required": panels_required,
            "Battery Capacity": f"{battery_capacity} kWh",
            "Daily Energy Required": f"{round(energy_required_per_day, 1)} kWh",
            "Daily Energy Production": f"{round(energy_production, 1)} kWh",
            "Specific Efficiency": f"{round(specific_efficiency, 2)} kg/kWh",
            
            # OUTPUTS - Production & Income
            "Daily Production": f"{round(production_per_day, 1)} kg",
            "Daily Gross Income": f"{selected_currency} {round(income_per_day * rate, 1)}",
            "Daily Net Income": f"{selected_currency} {round(net_income_per_day * rate, 1)}",
            "Annual Gross Income": f"{selected_currency} {round(gross_income_per_year * rate, 1)}",
            "Annual Net Profit": f"{selected_currency} {round(net_income_per_day * operating_days * rate, 1)}",
            
            # OUTPUTS - Cost Breakdown (in selected currency)
            "Machine Cost": f"{selected_currency} {round(price_usd * rate, 1)}",
            "Solar Panel Cost": f"{selected_currency} {round(solar_panel_cost * rate, 1)}",
            "Battery Cost": f"{selected_currency} {round(battery_cost * rate, 1)}",
            "Inverter/Controller Cost": f"{selected_currency} {round((inverter_cost + controller_cost) * rate, 1)}",
            "Import & Installation Cost": f"{selected_currency} {round((fob_subtotal_usd * (install_multiplier - 1)) * rate, 1)}",
            "FOB Subtotal": f"{selected_currency} {round(fob_subtotal_usd * rate, 1)}",
            "Total Installed Cost": f"{selected_currency} {round(total_with_import_usd * rate, 1)}",
            "Subsidy Amount": f"{selected_currency} {round(subsidy_amount * rate, 1)}",
            "Total After Subsidy": f"{selected_currency} {round(total_after_subsidy * rate, 1)}",
            "Deposit Amount": f"{selected_currency} {round(deposit_amount * rate, 1)}",
            "Loan Amount": f"{selected_currency} {round(loan_principal_usd * rate, 1)}",
            
            # OUTPUTS - Loan Details
            "Monthly Repayment": f"{selected_currency} {round(monthly_repayment_usd * rate, 1)}",
            "Annual Repayment": f"{selected_currency} {round(annual_repayment_usd * rate, 1)}",
            "Daily Repayment": f"{selected_currency} {round(daily_repayment_usd * rate, 1)}",
            "Total Interest Paid": f"{selected_currency} {round(total_interest_paid_usd * rate, 1)}",
            "Total Repayment Amount": f"{selected_currency} {round(total_repayment_usd * rate, 1)}",
            
            # OUTPUTS - Financial Analysis
            "Repayment % of Gross Revenue": f"{round(repayment_percentage, 1)}%",
            "Repayment % of Net Revenue": f"{round(net_revenue_repayment_percentage, 1)}%",
            "Payback Period": f"{round(payback_years_new, 1)} years" if viable_business else "N/A",
            "Business Viable": viability_text,
            "Daily Surplus": f"{selected_currency} {round((net_income_per_day - daily_repayment_usd) * rate, 1)}" if viable_business and net_income_per_day > daily_repayment_usd else f"{selected_currency} 0"
        }
        
        # --- Key Performance Indicators ---
        st.subheader("🎯 Key Performance Indicators")
        
        kpi_col1, kpi_col2, kpi_col3, kpi_col4 = st.columns(4)
        
        with kpi_col1:
            st.metric(
                "💰 Daily Net Income", 
                f"{selected_currency} {round(net_income_per_day * rate, 1)}",
                f"{selected_currency} {round((net_income_per_day - daily_repayment_usd) * rate, 1)} after loan"
            )
        
        with kpi_col2:
            st.metric(
                "🏗️ Total Investment", 
                f"{selected_currency} {round(total_after_subsidy * rate, 1)}",
                f"After {subsidy_percentage}% subsidy"
            )
        
        with kpi_col3:
            st.metric(
                "⏳ Payback Period", 
                f"{round(payback_years_new, 1)} years" if viable_business else "N/A",
                "Time to recover investment"
            )
        
        with kpi_col4:
            st.metric(
                "📈 Business Viability", 
                viability_text,
                "Based on cash flow analysis"
            )
        
        # --- Detailed Report Table ---
        st.divider()
        st.subheader("📋 Complete Analysis Report")
        
        # Convert to DataFrame for better display
        report_df = pd.DataFrame(list(report_data.items()), columns=["Parameter", "Value"])
        
        # Display in an expandable table
        with st.expander("📊 View Complete Report Data", expanded=True):
            st.dataframe(
                report_df,
                hide_index=True,
                use_container_width=True,
                height=600
            )
        
        # --- Export Section ---
        st.divider()
        st.subheader("📥 Export Options")
        
        col1, col2 = st.columns(2)
        
        with col1:
    # Enhanced XLSX Export with ACTUAL Excel formulas
            def create_excel_report():
                import io
                from openpyxl import Workbook
                from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
                from openpyxl.worksheet.datavalidation import DataValidation
                from openpyxl.styles import PatternFill, Font

                black_fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
                white_font = Font(color="FFFFFF", bold=True)


                
                # Create a workbook and select the active sheet
                wb = Workbook()
                ws = wb.active
                ws.title = "Solar Analysis Calculator"
                
                # Define styles
                header_font = Font(bold=True, size=14, color="FFFFFF")
                section_font = Font(bold=True, size=12, color="FFFFFF")
                label_font = Font(bold=True, size=10)
                value_font = Font(size=10)
                input_font = Font(size=10, color="2E7D32", bold=True)  # Green for inputs
                formula_font = Font(size=10, color="1976D2")  # Blue for formulas
                
                header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
                input_section_fill = PatternFill(start_color="E8F5E8", end_color="E8F5E8", fill_type="solid")
                formula_section_fill = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")
                output_section_fill = PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid")
                
                center_align = Alignment(horizontal='center', vertical='center')
                left_align = Alignment(horizontal='left', vertical='center')
                
                thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                                top=Side(style='thin'), bottom=Side(style='thin'))
                
                # Title and Header
                ws.merge_cells('A1:B1')
                ws['A1'] = "SOLAR PRODUCTIVE USE CALCULATOR - INTERACTIVE"
                ws['A1'].font = header_font
                ws['A1'].fill = header_fill
                ws['A1'].alignment = center_align
                
                ws.merge_cells('A2:B2')
                ws['A2'] = "🔧 Change values in GREEN cells and watch formulas recalculate automatically"
                ws['A2'].font = Font(italic=True, size=10, color="2E7D32")
                ws['A2'].alignment = center_align
                
                # Start data from row 4
                current_row = 4
                
                # SECTION 1: INPUT CELLS (Editable by users)
                ws.merge_cells(f'A{current_row}:B{current_row}')
                merged_range = ws[f'A{current_row}:B{current_row}']

                # Apply text and style to the first (anchor) cell
                cell = ws[f'A{current_row}']
                cell.value = "📥 INPUT PARAMETERS (Change these values)"
                cell.font = section_font
                cell.alignment = center_align
                cell.border = thin_border

                # Apply fill and border to all cells in the merged range
                for row in merged_range:
                    for c in row:
                        c.fill = input_section_fill
                        c.fill = black_fill
                        c.border = thin_border
                # Create dropdown for System Type (AC/DC)
                dv = DataValidation(type="list", formula1='"AC,DC"', allow_blank=False)
                dv.error = "Please select either AC or DC"
                dv.errorTitle = "Invalid Input"
                dv.prompt = "Select the system type"
                dv.promptTitle = "System Type"

                # Add the validation to the worksheet
                ws.add_data_validation(dv)
                dv.add("B5")

                # Optional: add label in column A
                ws["A5"] = "System Type (Choose from Dropdown)"

                current_row += 1
                
                # Define input cells with their current values
                # These will be editable by users
                input_cells = {
                    "B6": ("Power Consumption (kW)", power),
                    "B7": ("Processing Speed (kg/hour)", processing_speed),
                    "B8": ("Appliance Price (USD)", price_usd),
                    "B9": ("Daily Runtime (hours)", runtime_per_day),
                    "B10": ("Operating Days per Year", operating_days),
                    "B11": ("Income per kg (USD)", income_per_kg),
                    "B12": ("Sun Hours per Day", sun_hours),
                    "B13": ("System Efficiency (%)", system_efficiency),
                    "B14": ("Battery Storage (hours)", battery_hours),
                    "B15": ("Daily Operating Cost (USD)", daily_operating_cost),
                    "B16": ("Loan Term (Years)", loan_term_years),
                    "B17": ("Interest Rate (%)", interest_rate * 100),
                    "B18": ("Deposit (%)", deposit_percentage),
                    "B19": ("Import/Installation Cost Increase (%)", install_increase),
                    "B20": ("Subsidy (%)", subsidy_percentage),
                    "B21": ("Exchange Rate (USD to Local)", rate)
                }
                
                # Add input labels
                for cell_ref, (label, value) in input_cells.items():
                    row = int(cell_ref[1:])
                    ws[f'A{row}'] = label
                    ws[f'A{row}'].font = label_font
                    ws[f'A{row}'].alignment = left_align
                    ws[f'A{row}'].border = thin_border
                    
                    # Input cells are editable and colored
                    ws[cell_ref] = value
                    ws[cell_ref].font = input_font
                    ws[cell_ref].alignment = left_align
                    ws[cell_ref].border = thin_border
                    ws[cell_ref].fill = input_section_fill
                
                current_row = 22
                
                # SECTION 2: CALCULATIONS WITH ACTUAL EXCEL FORMULAS
                ws.merge_cells(f'A{current_row}:B{current_row}')
                ws[f'A{current_row}'] = "🧮 AUTOMATIC CALCULATIONS (These update automatically)"
                ws[f'A{current_row}'].font = section_font
                ws[f'A{current_row}'].fill = formula_section_fill
                ws[f'A{current_row}'].alignment = center_align
                ws[f'A{current_row}'].border = thin_border
                ws[f'A{current_row}'].fill = black_fill
                current_row += 1
                
                # Technical Calculations
                tech_calculations = {
                    f"A{current_row}": "Specific Efficiency (kg/kWh)",
                    f"B{current_row}": "=B7/B6",  # Processing Speed / Power
                    
                    f"A{current_row+1}": "Daily Energy Required (kWh)",
                    f"B{current_row+1}": "=B9*B6",  # Runtime * Power
                    
                    f"A{current_row+2}": "Energy Production Needed (kWh)",
                    f"B{current_row+2}": "=B24/(B13/100)",  # Energy Required / Efficiency
                    
                    f"A{current_row+3}": "Daily Production (kg)",
                    f"B{current_row+3}": "=B23*B24",  # Specific Efficiency * Energy Required
                    
                    f"A{current_row+4}": "Panel Energy per Day (kWh)",
                    f"B{current_row+4}": "=0.5*B12",  # Panel Wattage * Sun Hours
                    
                    f"A{current_row+5}": "Panels Required",
                    f"B{current_row+5}": "=ROUNDUP(B25 / (500 * B12 / 1000), 0)",  # CEILING(Energy Needed / Panel Energy)
                    
                    f"A{current_row+6}": "Recommended Solar Size (kWp)",
                    f"B{current_row+6}": "=CEILING(B25/B12,0.5)",  # Custom rounding logic
                    
                    f"A{current_row+7}": "Battery Capacity (kWh)",
                    f"B{current_row+7}": "=IFERROR((B29)*MAX(1,B14),"")",  # Solar Size * Battery Hours
                }
                
                current_row += 8
                
                # Cost Calculations
                cost_calculations = {
                    f"A{current_row}": "Solar Panel Cost (USD)",
                    f"B{current_row}": "=CEILING((B29*1000)/500,1)*50",  # Panels * $50 per panel
                    
                    f"A{current_row+1}": "Battery Cost (USD)",
                    f"B{current_row+1}": "=B30*300",  # Battery Capacity * $300/kWh
                    
                    f"A{current_row+2}": "Inverter/Controller Cost (USD)",
                    f"B{current_row+2}": '=IF(B5="AC",B29*100,B29*50)',  # Assuming AC system
                    
                    f"A{current_row+3}": "FOB Subtotal (USD)",
                    f"B{current_row+3}": "=B8+B31+B32+B33",  # Machine + Solar + Battery + Inverter
                    
                    f"A{current_row+4}": "Import & Installation Cost (USD)",
                    f"B{current_row+4}": "100%",  # FOB * Import %
                    
                    f"A{current_row+5}": "Total Installed Cost (USD)",
                    f"B{current_row+5}": "=B34*(1+B35)",  # FOB + Import
                    
                    f"A{current_row+6}": "Subsidy Amount (USD)",
                    f"B{current_row+6}": "=B36*(B20/100)",  # Total Cost * Subsidy %
                    
                    f"A{current_row+7}": "Total After Subsidy (USD)",
                    f"B{current_row+7}": "=B36-B37",  # Total - Subsidy
                    
                    f"A{current_row+8}": "Deposit Amount (USD)",
                    f"B{current_row+8}": "=B38*(B18/100)",  # After Subsidy * Deposit %
                    
                    f"A{current_row+9}": "Loan Principal (USD)",
                    f"B{current_row+9}": "=B38-B39",  # After Subsidy - Deposit
                }
                
                current_row += 10
                
                # Income & Loan Calculations
                income_calculations = {
                    f"A{current_row}": "Daily Gross Income (USD)",
                    f"B{current_row}": "=B11*B26*B21",  # Income per kg * Daily Production
                    
                    f"A{current_row+1}": "Daily Net Income (USD)",
                    f"B{current_row+1}": "=(B41/B21-B15)*B21",  # Gross Income - Operating Cost
                    
                    f"A{current_row+2}": "Annual Gross Profit (USD)",
                    f"B{current_row+2}": "=B41*B10*B21",  # Daily Gross * Operating Days

                    f"A{current_row+2}": "Annual Net Profit (USD)",
                    f"B{current_row+2}": "=B42*B10*B21",  # Daily Net * Operating Days
                    
                    f"A{current_row+3}": "Monthly Interest Rate",
                    f"B{current_row+3}": "=B17/12/100",  # Annual Rate / 12 months
                    
                    f"A{current_row+4}": "Monthly Repayment (USD)",
                    f"B{current_row+4}": "=(B40*B44)/(1-(1+B44)^(-B16*12))",  # Loan payment formula
                    
                    f"A{current_row+5}": "Annual Repayment (USD)",
                    f"B{current_row+5}": "=B45*12",  # Monthly * 12
                    
                    f"A{current_row+6}": "Daily Repayment (USD)",
                    f"B{current_row+6}": "=B46/365",  # Annual / 365
                    
                    f"A{current_row+7}": "Total Repayment Amount (USD)",
                    f"B{current_row+7}": "=B45*B16*12",  # Monthly * Months
                    
                    f"A{current_row+8}": "Total Interest Paid (USD)",
                    f"B{current_row+8}": "=B48-B40",  # Total Repayment - Principal
                    
                    f"A{current_row+9}": "Payback Period (Years)",
                    f"B{current_row+9}": "=IF(B61>0, B38/B43, \"N/A\")",  # Investment / Annual Profit

                    f"A{current_row+10}": "Current Business Viability",
                    f"B{current_row+10}": '=IF(OR(B42="",B47=""),"",IF(B42>=B47,"Yes ✅","No ❌"))',

                    
                }
                
                # Apply all formulas to the worksheet
                all_calculations = {**tech_calculations, **cost_calculations, **income_calculations}
                
                for cell_ref, value in all_calculations.items():
                    if cell_ref.startswith('A'):
                        # This is a label
                        ws[cell_ref] = value
                        ws[cell_ref].font = label_font
                        ws[cell_ref].alignment = left_align
                        ws[cell_ref].border = thin_border
                    else:
                        # This is a formula or value
                        if isinstance(value, str) and value.startswith('='):
                            # It's a formula
                            ws[cell_ref] = value
                            ws[cell_ref].font = formula_font
                        else:
                            # It's a value
                            ws[cell_ref] = value
                            ws[cell_ref].font = value_font
                        
                        ws[cell_ref].alignment = left_align
                        ws[cell_ref].border = thin_border
                        ws[cell_ref].fill = formula_section_fill
                
                # Final row for current results display
                current_row += 15
                ws.merge_cells(f'A{current_row}:B{current_row}')
                ws[f'A{current_row}'] = "💡 Prefilled Results Based on Your Inputs through Online Calculator"
                ws[f'A{current_row}'].font = Font(bold=True, size=11, color="2C3E50")
                ws[f'A{current_row}'].alignment = center_align
                current_row += 1
                
                # Display current calculated values
                current_results = {
                    f"A{current_row}": "Machine Name",
                    f"B{current_row}": f"{selected_appliance}",

                    f"A{current_row+1}": "System Rating",
                    f"B{current_row+1}": f"{selected_system}",

                    f"A{current_row+2}": "Current Solar System Size",
                    f"B{current_row+2}": f"{recommended_solar_size} kWp",
                    
                    f"A{current_row+3}": "Current Daily Net Income",
                    f"B{current_row+3}": f"{selected_currency} {round(net_income_per_day * rate, 1)}",
                    
                    f"A{current_row+4}": "Current Total Investment", 
                    f"B{current_row+4}": f"{selected_currency} {round(total_after_subsidy * rate, 1)}",
                    
                    f"A{current_row+5}": "Current Payback Period",
                    f"B{current_row+5}": f"{round(payback_years_new, 1)} years" if viable_business else "N/A",
                    
                    f"A{current_row+6}": "Current Business Viability",
                    f"B{current_row+6}": viability_text,
                }
                
                for cell_ref, value in current_results.items():
                    ws[cell_ref] = value
                    if cell_ref.startswith('A'):
                        ws[cell_ref].font = label_font
                    else:
                        ws[cell_ref].font = value_font
                    ws[cell_ref].alignment = left_align
                    ws[cell_ref].border = thin_border
                
                # Set column widths
                ws.column_dimensions['A'].width = 35
                ws.column_dimensions['B'].width = 25
                
                # Save to bytes
                excel_buffer = io.BytesIO()
                wb.save(excel_buffer)
                excel_buffer.seek(0)
                
                return excel_buffer.getvalue()
            
            # Create dropdown for export format
            export_format = st.selectbox(
                "Choose Export Format:",
                ["Select format", "Excel (.xlsx) - Interactive Calculator", "CSV (.csv) - Simple Report"],
                key="export_format"
            )
            
            def create_simple_csv():
                """Create a simple CSV version"""
                csv_data = []
                csv_data.append("SOLAR PRODUCTIVE USE CALCULATOR REPORT")
                csv_data.append(f"Generated: {report_data['Date']}")
                csv_data.append(f"Currency: {selected_currency}")
                csv_data.append("")
                
                # Inputs
                csv_data.append("INPUT PARAMETERS")
                csv_data.append("Parameter,Value")
                input_params = {
                    "Appliance": selected_appliance,
                    "System Type": selected_system,
                    "Power (kW)": power,
                    "Processing Speed (kg/hr)": processing_speed,
                    "Daily Runtime (hrs)": runtime_per_day,
                    "Operating Days/Year": operating_days,
                    "Sun Hours/Day": sun_hours,
                    "System Efficiency (%)": system_efficiency,
                    "Income per kg": f"{selected_currency} {round(income_per_kg * rate, 3)}",
                    "Daily Operating Cost": f"{selected_currency} {round(daily_operating_cost * rate, 1)}"
                }
                
                for key, value in input_params.items():
                    csv_data.append(f"{key},{value}")
                
                csv_data.append("")
                
                # Results
                csv_data.append("CALCULATED RESULTS")
                csv_data.append("Parameter,Value")
                results = {
                    "Solar System Size": f"{recommended_solar_size} kWp",
                    "Panels Required": f"{panels_required}",
                    "Daily Production": f"{round(production_per_day, 1)} kg",
                    "Daily Net Income": f"{selected_currency} {round(net_income_per_day * rate, 1)}",
                    "Total Investment": f"{selected_currency} {round(total_after_subsidy * rate, 1)}",
                    "Payback Period": f"{round(payback_years_new, 1)} years" if viable_business else "N/A",
                    "Business Viable": viability_text
                }
                
                for key, value in results.items():
                    csv_data.append(f"{key},{value}")
                
                return "\n".join(csv_data)
            
            # Download button based on selection
            if export_format == "Excel (.xlsx) - Interactive Calculator":
                if st.button("📊 Generate Interactive Excel", use_container_width=True):
                    with st.spinner("Creating interactive Excel calculator..."):
                        try:
                            excel_data = create_excel_report()
                            st.download_button(
                                label="📥 Download Interactive Excel (.xlsx)",
                                data=excel_data,
                                file_name=f"solar_calculator_interactive_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                use_container_width=True,
                                help="Change GREEN input cells and watch formulas recalculate automatically!"
                            )
                        except Exception as e:
                            st.error(f"Error generating Excel report: {e}")
                            st.info("Make sure you have openpyxl installed: pip install openpyxl")
            
            elif export_format == "CSV (.csv) - Simple Report":
                if st.button("📄 Generate CSV Report", use_container_width=True):
                    with st.spinner("Generating CSV report..."):
                        try:
                            csv_data = create_simple_csv()
                            st.download_button(
                                label="📥 Download CSV Report (.csv)",
                                data=csv_data,
                                file_name=f"solar_calculator_report_{datetime.now().strftime('%Y%m%d_%H%M')}.csv",
                                mime="text/csv",
                                use_container_width=True,
                                help="Simple CSV format for quick viewing"
                            )
                        except Exception as e:
                            st.error(f"Error generating CSV report: {e}")

        with col2:
            # Enhanced PDF Export with table format
            from reportlab.lib.pagesizes import letter
            from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.enums import TA_CENTER, TA_LEFT
            from reportlab.lib import colors
            import tempfile
            import requests
            from io import BytesIO
            
            def create_comprehensive_pdf():
                with tempfile.NamedTemporaryFile(delete=False, suffix=".pdf") as tmp:
                    pdf_path = tmp.name
                
                doc = SimpleDocTemplate(pdf_path, pagesize=letter)
                styles = getSampleStyleSheet()
                
                # Custom styles
                title_style = ParagraphStyle("TitleCenter", parent=styles["Title"], alignment=TA_CENTER, fontSize=16, spaceAfter=20)
                section_style = ParagraphStyle("Section", parent=styles["Heading2"], fontSize=12, spaceAfter=12, textColor=colors.HexColor("#2c3e50"))
                normal_style = ParagraphStyle("Normal", parent=styles["Normal"], alignment=TA_LEFT, leading=12)
                
                story = []
                
                # Title and Header
                story.append(Paragraph("SOLAR PRODUCTIVE USE CALCULATOR", title_style))
                story.append(Paragraph("COMPREHENSIVE ANALYSIS REPORT", title_style))
                story.append(Spacer(1, 10))
                story.append(Paragraph(f"Generated on: {report_data['Date']}", styles["Italic"]))
                story.append(Spacer(1, 20))
                
                # Appliance Image
                appliance_images = {
                    "Hammer Mill for Flour": "https://productivesolarsolutions.com/uploads/products/MaizeMill.png",
                    "Rice Huller/Polisher": "https://productivesolarsolutions.com/uploads/products/RiceMill-MaizeHuller.png"
                }
                
                if selected_appliance in appliance_images:
                    try:
                        response = requests.get(appliance_images[selected_appliance])
                        img_data = BytesIO(response.content)
                        img = Image(img_data, width=180, height=135)
                        img.hAlign = 'CENTER'
                        story.append(img)
                        story.append(Spacer(1, 8))
                        story.append(Paragraph(f"<i>{selected_appliance}</i>", styles["Italic"]))
                        story.append(Spacer(1, 20))
                    except:
                        pass
                
                # Key Performance Indicators Table
                story.append(Paragraph("KEY PERFORMANCE INDICATORS", section_style))
                
                kpi_data = [
                    ["DAILY NET INCOME", report_data["Daily Net Income"]],
                    ["TOTAL INVESTMENT", report_data["Total After Subsidy"]],
                    ["PAYBACK PERIOD", report_data["Payback Period"]],
                    ["BUSINESS VIABILITY", report_data["Business Viable"]],
                    ["SOLAR SYSTEM SIZE", report_data["Recommended Solar Size"]],
                    ["DAILY PRODUCTION", report_data["Daily Production"]]
                ]
                
                kpi_table = Table(kpi_data, colWidths=[200, 200])
                kpi_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor("#f8f9fa")),
                    ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, -1), 10),
                    ('GRID', (0, 0), (-1, -1), 1, colors.grey),
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#4CAF50")),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ]))
                story.append(kpi_table)
                story.append(Spacer(1, 25))
                
                # Input Parameters Table
                story.append(Paragraph("INPUT PARAMETERS", section_style))
                
                input_data = [
                    ["Parameter", "Value"],
                    ["Appliance", report_data["Appliance"]],
                    ["System Type", report_data["System Type"]],
                    ["Power Consumption", report_data["Power Consumption"]],
                    ["Processing Speed", report_data["Processing Speed"]],
                    ["Appliance Price", report_data["Appliance Price"]],
                    ["Daily Runtime", report_data["Daily Runtime"]],
                    ["Operating Days/Year", report_data["Operating Days/Year"]],
                    ["Sun Hours/Day", report_data["Sun Hours/Day"]],
                    ["System Efficiency", report_data["System Efficiency"]],
                    ["Battery Storage", report_data["Battery Storage"]],
                    ["Income per kg", report_data["Income per kg"]],
                    ["Daily Operating Cost", report_data["Daily Operating Cost"]],
                    ["Loan Term", report_data["Loan Term"]],
                    ["Interest Rate", report_data["Interest Rate"]],
                    ["Deposit Percentage", report_data["Deposit Percentage"]],
                    ["Import/Installation Cost Increase", report_data["Import/Installation Cost Increase"]],
                    ["Subsidy Percentage", report_data["Subsidy Percentage"]]
                ]
                
                input_table = Table(input_data, colWidths=[220, 220])
                input_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#2c3e50")),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                    ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, -1), 9),
                    ('GRID', (0, 0), (-1, -1), 1, colors.grey),
                    ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor("#ffffff")),
                ]))
                story.append(input_table)
                story.append(Spacer(1, 25))
                
                # Technical Specifications Table
                story.append(Paragraph("TECHNICAL SPECIFICATIONS", section_style))
                
                tech_data = [
                    ["Parameter", "Value"],
                    ["Recommended Solar Size", report_data["Recommended Solar Size"]],
                    ["Panels Required", report_data["Panels Required"]],
                    ["Battery Capacity", report_data["Battery Capacity"]],
                    ["Daily Energy Required", report_data["Daily Energy Required"]],
                    ["Daily Energy Production", report_data["Daily Energy Production"]],
                    ["Specific Efficiency", report_data["Specific Efficiency"]],
                    ["Daily Production", report_data["Daily Production"]]
                ]
                
                tech_table = Table(tech_data, colWidths=[220, 220])
                tech_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#3498db")),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                    ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, -1), 9),
                    ('GRID', (0, 0), (-1, -1), 1, colors.grey),
                    ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor("#f0f8ff")),
                ]))
                story.append(tech_table)
                story.append(Spacer(1, 25))
                
                # Financial Analysis Table
                story.append(Paragraph("FINANCIAL ANALYSIS", section_style))
                
                financial_data = [
                    ["Parameter", "Value"],
                    ["Daily Gross Income", report_data["Daily Gross Income"]],
                    ["Daily Net Income", report_data["Daily Net Income"]],
                    ["Annual Gross Income", report_data["Annual Gross Income"]],
                    ["Annual Net Profit", report_data["Annual Net Profit"]],
                    ["Machine Cost", report_data["Machine Cost"]],
                    ["Solar Panel Cost", report_data["Solar Panel Cost"]],
                    ["Battery Cost", report_data["Battery Cost"]],
                    ["Inverter/Controller Cost", report_data["Inverter/Controller Cost"]],
                    ["Import & Installation Cost", report_data["Import & Installation Cost"]],
                    ["FOB Subtotal", report_data["FOB Subtotal"]],
                    ["Total Installed Cost", report_data["Total Installed Cost"]],
                    ["Subsidy Amount", report_data["Subsidy Amount"]],
                    ["Total After Subsidy", report_data["Total After Subsidy"]],
                    ["Deposit Amount", report_data["Deposit Amount"]],
                    ["Loan Amount", report_data["Loan Amount"]]
                ]
                
                financial_table = Table(financial_data, colWidths=[220, 220])
                financial_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#27ae60")),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                    ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, -1), 9),
                    ('GRID', (0, 0), (-1, -1), 1, colors.grey),
                    ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor("#f0fff4")),
                ]))
                story.append(financial_table)
                story.append(Spacer(1, 25))
                
                # Loan Analysis Table
                story.append(Paragraph("LOAN ANALYSIS", section_style))
                
                loan_data = [
                    ["Parameter", "Value"],
                    ["Monthly Repayment", report_data["Monthly Repayment"]],
                    ["Annual Repayment", report_data["Annual Repayment"]],
                    ["Daily Repayment", report_data["Daily Repayment"]],
                    ["Total Interest Paid", report_data["Total Interest Paid"]],
                    ["Total Repayment Amount", report_data["Total Repayment Amount"]],
                    ["Repayment % of Gross Revenue", report_data["Repayment % of Gross Revenue"]],
                    ["Repayment % of Net Revenue", report_data["Repayment % of Net Revenue"]],
                    ["Payback Period", report_data["Payback Period"]],
                    ["Business Viable", report_data["Business Viable"]],
                    ["Daily Surplus", report_data["Daily Surplus"]]
                ]
                
                loan_table = Table(loan_data, colWidths=[220, 220])
                loan_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#e74c3c")),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                    ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, -1), 9),
                    ('GRID', (0, 0), (-1, -1), 1, colors.grey),
                    ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor("#fff5f5")),
                ]))
                story.append(loan_table)
                story.append(Spacer(1, 20))
                
                # Footer
                story.append(Paragraph("--- End of Report ---", styles["Italic"]))
                story.append(Spacer(1, 10))
                story.append(Paragraph(f"Currency: {selected_currency} | Exchange Rate: 1 USD = {rate:.2f} {selected_currency}", styles["Italic"]))
                
                # Build PDF
                doc.build(story)
                
                with open(pdf_path, 'rb') as f:
                    return f.read()
            
            if st.button("📄 Generate Comprehensive PDF Report", use_container_width=True):
                with st.spinner("Generating detailed PDF report..."):
                    try:
                        pdf_bytes = create_comprehensive_pdf()
                        st.download_button(
                            label="📥 Download PDF Report",
                            data=pdf_bytes,
                            file_name=f"solar_calculator_comprehensive_report_{datetime.now().strftime('%Y%m%d_%H%M')}.pdf",
                            mime="application/pdf",
                            use_container_width=True
                        )
                    except Exception as e:
                        st.error(f"Error generating PDF: {e}")
                        st.info("Make sure you have reportlab installed: pip install reportlab")

    
if st.sidebar.checkbox("📦 Manage Appliances"):
    st.sidebar.markdown("---")
    st.sidebar.subheader("➕ Add or Edit Appliances")

    # Use radio buttons instead of tabs (since sidebar doesn’t support tabs)
    with st.sidebar:
        st.markdown("## 🔧 Admin Panel")

        # --- Review Submissions Section ---
        with st.expander("📋 Review Submissions"):
            if not st.session_state.pending_submissions:
                st.info("No pending submissions")
            else:
                for submission in st.session_state.pending_submissions:
                    with st.expander(f"{submission['name']}"):
                        st.write(f"**AC Power:** {submission['power_ac']} kW")
                        st.write(f"**DC Power:** {submission['power_dc']} kW")
                        st.write(f"**Processing Speed:** {submission['processing_speed']} kg/hr")
                        st.write(f"**Price:** ${submission['price_usd']}")
                        st.write(f"**Description:** {submission['description']}")
                        st.write(f"**Submitted:** {submission['submission_date'][:10]}")

                        col1, col2 = st.columns(2)
                        with col1:
                            if st.button("✅ Approve", key=f"approve_{submission['id']}"):
                                approved_appliance = {
                                    "name": submission["name"],
                                    "power_ac": submission["power_ac"],
                                    "power_dc": submission["power_dc"],
                                    "processing_speed": submission["processing_speed"],
                                    "price_usd": submission["price_usd"]
                                }
                                st.session_state.custom_appliances.append(approved_appliance)
                                save_custom_appliances(st.session_state.custom_appliances)

                                # update maps
                                power_map[approved_appliance["name"]] = approved_appliance["power_ac"]
                                power_map_dc[approved_appliance["name"]] = approved_appliance["power_dc"]

                                # remove from pending
                                st.session_state.pending_submissions = [
                                    s for s in st.session_state.pending_submissions if s["id"] != submission["id"]
                                ]
                                save_pending_submissions(st.session_state.pending_submissions)
                                st.rerun()

                        with col2:
                            if st.button("❌ Reject", key=f"reject_{submission['id']}"):
                                st.session_state.pending_submissions = [
                                    s for s in st.session_state.pending_submissions if s["id"] != submission["id"]
                                ]
                                save_pending_submissions(st.session_state.pending_submissions)
                                st.rerun()

        # --- Manage Appliances Section ---
        with st.expander("⚙️ Manage Appliances"):
            action = st.radio("Choose Action", ["➕ Add New", "✏️ Edit/Delete Existing"], key="manage_action")

            if action == "➕ Add New":
                custom_appliance_submission_form()

            elif action == "✏️ Edit/Delete Existing":
                if not st.session_state.custom_appliances:
                    st.info("No appliances to edit or delete.")
                else:
                    appliance_names = [app["name"] for app in st.session_state.custom_appliances]
                    selected_name = st.selectbox("Select Appliance", appliance_names, key="edit_select")

                    appliance = next(app for app in st.session_state.custom_appliances if app["name"] == selected_name)

                    # Editable form
                    with st.form(f"edit_form_{selected_name}"):
                        col1, col2 = st.columns(2)
                        with col1:
                            name = st.text_input("Appliance Name*", value=appliance["name"])
                            power_ac = st.number_input("AC Power (kW)", min_value=0.1, value=float(appliance["power_ac"]), step=0.1)
                            power_dc = st.number_input("DC Power (kW)", min_value=0.1, value=float(appliance["power_dc"]), step=0.1)
                            processing_speed = st.number_input("Processing Speed (kg/hr)", min_value=1, value=int(appliance["processing_speed"]), step=1)
                        with col2:
                            price_usd = st.number_input("Price (USD)", min_value=0.0, value=float(appliance["price_usd"]), step=50.0)
                            description = st.text_area("Description", value=appliance.get("description", ""))
                            contact_email = st.text_input("Contact Email", value=appliance.get("contact_email", ""))

                        update_btn = st.form_submit_button("💾 Update Appliance")

                    if update_btn:
                        appliance.update({
                            "name": name,
                            "power_ac": power_ac,
                            "power_dc": power_dc,
                            "processing_speed": processing_speed,
                            "price_usd": price_usd,
                            "description": description,
                            "contact_email": contact_email,
                            "last_updated": datetime.now().isoformat()
                        })
                        save_custom_appliances(st.session_state.custom_appliances)
                        save_custom_appliance_to_csv(appliance)
                        st.success(f"✅ {name} updated successfully!")

                    # Delete button (separate from form to avoid conflicts)
                    if st.button(f"🗑️ Delete {selected_name}", key=f"delete_{selected_name}"):
                        st.session_state.custom_appliances = [
                            app for app in st.session_state.custom_appliances if app["name"] != selected_name
                        ]
                        save_custom_appliances(st.session_state.custom_appliances)
                        st.warning(f"🗑️ {selected_name} deleted successfully!")
                        st.rerun()



    
